#coding=gbk

from PIL import Image, ImageTk
import tkinter
import tkinter.font as tkfront
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import sys
import findindex
import interface

iface = interface.Interface()
wb = openpyxl.load_workbook(interface.workbook)
ws0 = wb[wb.sheetnames[0]]
ws = wb[wb.sheetnames[1]]
ret_dict = findindex.find(ws0, ws, interface.follow_names, interface.follow_cols)
follow_line_list = ret_dict["follow_line"]
follow_line_list.insert(0, interface.first_row)
pic_coord_list_dict = ret_dict["follow_index"]


tk = tkinter.Tk()

tk.title(interface.title)
tk.geometry(interface.geometry)
ft = tkfront.Font(size=interface.front_size)
img_loader = SheetImageLoader(ws)

def LoadLabelss():
	local_row = 1
	for row in follow_line_list:
		first_cell = ws0.cell(row, 1)
		if None == first_cell:
			continue
		is_filter_line = False
		index = first_cell.value
		local_col = 1
		labels = []
		for col in interface.display_cols:
			cell = ws0.cell(row, col)
			if None == cell:
				continue
			str = cell.value
			if str == None:
				str = " "
			label = tkinter.Label(tk, text = str, font=ft, wraplength=interface.wraplength)
			label.grid(column=local_col, row=local_row, sticky='w')
			if local_row > interface.one_page_size + 1:
				label.grid_remove()
			labels.append(label)
			if col == interface.filter_col and str == " ":
				is_filter_line = True
			local_col += 1

		if local_row > 1:
			pic_coord_list = pic_coord_list_dict[index]
			for i in range(len(pic_coord_list)):
				pic_coord = pic_coord_list[i]
				if False == img_loader.image_in(pic_coord):
					continue
				img = img_loader.get(pic_coord)
				tk_img = ImageTk.PhotoImage(img)
				label = tkinter.Label(tk, image = tk_img)
				label.image = tk_img
				label.grid(column=i + 1, row=interface.third_row)
				label.grid_remove()
				labels.append(label)
				local_col += 1
			if True == is_filter_line:
				iface.filter_labelss.append(labels)
			iface.labelss.append(labels)
			if local_row <= interface.one_page_size + 1:
				iface.grid_labels_index.append(local_row - 2)
		else:
			iface.title_labels = labels
			for i in range(len(labels)):
				label = labels[i]
				temp_label = tkinter.Label(tk, text = label.cget("text"), font=ft, wraplength=interface.wraplength)
				temp_label.grid(column=i + 1, row=interface.second_row, sticky='w')
				temp_label.grid_remove()
				iface.temporary_labels.append(temp_label)
		local_row += 1



LoadLabelss()
iface.Display()

tk.bind("<Key>", lambda event: iface.key(event, tk))

wb.close()
tk.mainloop()
